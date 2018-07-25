# Python Script to read and copy the source excel into de...
# Python Script to read and copy the source excel into destination excel sheet
# The destination columns have been hardcoded
 
import openpyxl
from   openpyxl import load_workbook
import os                                     # Macro enabling 
#wbname='IT.xlsm'    #Destination file
#Prepare the spreadsheets to copy from and paste too.
 
#File to be copied(Input file name)
wb = openpyxl.load_workbook("incident_Sample.xlsx") #Add  Source Excel file name
sheet = wb.get_sheet_by_name("Page 1") #Add Source Excel  Sheet name
row_count = sheet.max_row
 
#File to be pasted into  (Input the output file)

template = openpyxl.load_workbook("h1.xlsx")
template = openpyxl.load_workbook("IT.xlsm", keep_vba=True) #Add destination  Excel  file name     # Macro Preservation Using keep_vba
temp_sheet = template.get_sheet_by_name("Page 1") #Add destination excel Sheet name
row_count2 = temp_sheet.max_row

end_row1 = row_count + row_count2
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow + 1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
  
def createData():
    print("Processing...")
    selectedRange = copyRange(1,2,1,row_count,sheet) #Change the 4 number values as per the requirement. First parameter is column.
    pastingRange = pasteRange(1,(row_count2 + 1),1,(end_row1 - 1),temp_sheet,selectedRange) #Change the 4 number values

    selectedRange = copyRange(2,2,2,row_count,sheet)
    pastingRange = pasteRange(9,(row_count2 + 1),9,(end_row1 - 1),temp_sheet,selectedRange)

    selectedRange = copyRange(3,2,3,row_count,sheet)
    pastingRange = pasteRange(13,(row_count2 + 1),13,(end_row1 - 1),temp_sheet,selectedRange)

    selectedRange = copyRange(4,2,4,row_count,sheet)
    pastingRange = pasteRange(3,(row_count2 + 1),3,(end_row1 - 1),temp_sheet,selectedRange)

    selectedRange = copyRange(5,2,5,row_count,sheet)
    pastingRange = pasteRange(4,(row_count2 + 1),4,(end_row1 - 1),temp_sheet,selectedRange)

    selectedRange = copyRange(6,2,6,row_count,sheet)
    pastingRange = pasteRange(8,(row_count2 + 1),8,(end_row1 - 1),temp_sheet,selectedRange)

    selectedRange = copyRange(7,2,7,row_count,sheet)
    pastingRange = pasteRange(10,(row_count2 + 1),10,(end_row1 - 1),temp_sheet,selectedRange)
    
    selectedRange = copyRange(8,2,8,row_count,sheet)
    pastingRange = pasteRange(10,(row_count2 + 1),10,(end_row1 - 1),temp_sheet,selectedRange)

    selectedRange = copyRange(9,2,9,row_count,sheet)
    pastingRange = pasteRange(14,(row_count2 + 1),14,(end_row1 - 1),temp_sheet,selectedRange)

    selectedRange = copyRange(10,2,10,row_count,sheet)
    pastingRange = pasteRange(15,(row_count2 + 1),15,(end_row1 - 1),temp_sheet,selectedRange)

    selectedRange = copyRange(11,2,11,row_count,sheet)
    pastingRange = pasteRange(17,(row_count2 + 1),17,(end_row1 - 1),temp_sheet,selectedRange)

    #You can save the template as another file to create a new file here too.s
    template.save("h1.xlsx")           # Save the destination file to the temp file
   # os.rename('temp.xlsm', wbname)       # Rename the file to the original file name
    print("Range copied and pasted!")
go = createData()

